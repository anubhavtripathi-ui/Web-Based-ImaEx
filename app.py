import streamlit as st
import pytesseract
from PIL import Image
import pandas as pd
import re
import io

st.set_page_config(
    page_title="Mobile Number Extractor",
    page_icon="📱",
    layout="wide"
)

st.title("📱 Mobile Number Extractor")
st.write("Upload up to **30 images** at once. The app will extract all 10-digit mobile numbers and calculate their digital root sum.")


def digital_root(number_str):
    total = sum(int(d) for d in number_str)
    while total >= 10:
        total = sum(int(d) for d in str(total))
    return total


def extract_numbers_from_image(image):
    configs = [
        "--oem 3 --psm 6",
        "--oem 3 --psm 4",
        "--oem 3 --psm 11",
        "--oem 3 --psm 12",
    ]
    all_numbers = []
    seen = set()

    for config in configs:
        try:
            text = pytesseract.image_to_string(image, config=config)
            cleaned = re.sub(r"[^\d\s\n]", "", text)
            found = re.findall(r"\b[6-9]\d{9}\b", cleaned)
            for num in found:
                if num not in seen:
                    seen.add(num)
                    all_numbers.append(num)
        except Exception:
            continue

    if not all_numbers:
        for config in configs:
            try:
                text = pytesseract.image_to_string(image, config=config)
                found = re.findall(r"\d{10}", text)
                for num in found:
                    if num not in seen:
                        seen.add(num)
                        all_numbers.append(num)
            except Exception:
                continue

    return all_numbers


def get_low_confidence_numbers(image, numbers):
    low_conf = set()
    try:
        data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
        for i, word in enumerate(data["text"]):
            cleaned = re.sub(r"\D", "", word)
            if len(cleaned) >= 4:
                conf = int(data["conf"][i])
                if conf < 35:
                    for num in numbers:
                        if cleaned in num:
                            low_conf.add(num)
    except Exception:
        pass
    return low_conf


def style_excel(writer, df):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    workbook = writer.book
    worksheet = writer.sheets["Mobile Numbers"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="AAAAAA")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border

    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    na_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    na_font = Font(color="C00000", bold=True)

    sum_col_idx = df.columns.get_loc("Digital Root Sum") + 1

    for row_num in range(2, len(df) + 2):
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border
            if row_num % 2 == 0:
                cell.fill = alt_fill

        sum_cell = worksheet.cell(row=row_num, column=sum_col_idx)
        if sum_cell.value == "NA":
            sum_cell.fill = na_fill
            sum_cell.font = na_font

    col_widths = {"Sr. No.": 10, "Source Image": 30, "Mobile Number": 22, "Digital Root Sum": 20}
    for col_name, width in col_widths.items():
        if col_name in df.columns:
            col_letter = chr(ord("A") + df.columns.get_loc(col_name))
            worksheet.column_dimensions[col_letter].width = width

    worksheet.row_dimensions[1].height = 22


uploaded_files = st.file_uploader(
    "Upload Images (max 30)",
    type=["png", "jpg", "jpeg", "bmp", "tiff", "webp"],
    accept_multiple_files=True,
    help="Select up to 30 images at once — PNG, JPG, JPEG, BMP, TIFF, WebP"
)

if uploaded_files:
    if len(uploaded_files) > 30:
        st.error("Please upload a maximum of 30 images at a time.")
        st.stop()

    st.info(f"**{len(uploaded_files)}** image(s) selected. Processing...")

    all_rows = []
    global_sr = 1
    per_image_results = []

    progress = st.progress(0, text="Starting...")

    for file_idx, uploaded_file in enumerate(uploaded_files):
        progress.progress(
            (file_idx) / len(uploaded_files),
            text=f"Processing image {file_idx + 1} of {len(uploaded_files)}: {uploaded_file.name}"
        )

        try:
            image = Image.open(uploaded_file)
            numbers = extract_numbers_from_image(image)
            low_conf = get_low_confidence_numbers(image, numbers)

            rows_for_image = []
            for num in numbers:
                if num in low_conf:
                    rows_for_image.append({
                        "Sr. No.": global_sr,
                        "Source Image": uploaded_file.name,
                        "Mobile Number": num,
                        "Digital Root Sum": "NA",
                    })
                else:
                    rows_for_image.append({
                        "Sr. No.": global_sr,
                        "Source Image": uploaded_file.name,
                        "Mobile Number": num,
                        "Digital Root Sum": digital_root(num),
                    })
                global_sr += 1
                all_rows.append(rows_for_image[-1])

            per_image_results.append({
                "file": uploaded_file.name,
                "image": image,
                "count": len(numbers),
                "rows": rows_for_image,
                "error": None,
            })

        except Exception as e:
            per_image_results.append({
                "file": uploaded_file.name,
                "image": None,
                "count": 0,
                "rows": [],
                "error": str(e),
            })

    progress.progress(1.0, text="Done!")

    st.divider()

    total_found = sum(r["count"] for r in per_image_results)
    errors = [r for r in per_image_results if r["error"]]
    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Images Processed", len(uploaded_files))
    col_b.metric("Total Numbers Found", total_found)
    col_c.metric("Errors", len(errors))

    if all_rows:
        st.subheader("📊 All Extracted Numbers")
        df = pd.DataFrame(all_rows)
        st.dataframe(df, use_container_width=True, hide_index=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Mobile Numbers")
            style_excel(writer, df)
        output.seek(0)

        st.download_button(
            label="⬇️ Download Excel File",
            data=output,
            file_name="mobile_numbers.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        na_count = sum(1 for r in all_rows if r["Digital Root Sum"] == "NA")
        if na_count > 0:
            st.info(f"ℹ️ {na_count} number(s) were marked **NA** due to low OCR confidence.")
    else:
        st.warning("No 10-digit mobile numbers found in any of the uploaded images.")

    st.divider()
    st.subheader("Per-Image Results")
    for result in per_image_results:
        with st.expander(f"{'✅' if not result['error'] else '❌'} {result['file']} — {result['count']} number(s) found"):
            if result["error"]:
                st.error(f"Error: {result['error']}")
            elif result["image"] is not None:
                col1, col2 = st.columns([1, 2])
                with col1:
                    st.image(result["image"], use_column_width=True)
                with col2:
                    if result["rows"]:
                        st.dataframe(
                            pd.DataFrame(result["rows"])[["Mobile Number", "Digital Root Sum"]],
                            use_container_width=True,
                            hide_index=True,
                        )
                    else:
                        st.write("No numbers found.")

else:
    st.info("Upload one or more images to get started.")

st.divider()
st.caption("All processing is done locally — no data is sent to any server.")
